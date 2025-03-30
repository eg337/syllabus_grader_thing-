from openpyxl import Workbook


def write_calc(grade_dict):
    COLS = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
    row = 2
    categories = set()
    for key in grade_dict:
        if key != "course title":
            categories.add(key)
    # Create a new workbook
    workbook = Workbook()

    # Get the active worksheet
    worksheet = workbook.active

    # Add data to cells
    worksheet['A1'] = 'Course'
    worksheet['B1'] = 'Name'
    worksheet['C1'] = 'Deadline'
    worksheet['D1'] = 'Category'
    worksheet['E1'] = 'Weight'
    worksheet['F1'] = 'Score'



    worksheet['H1'] = "Course"
    worksheet['I1'] = "Grade"

    title = grade_dict['course title']
    deadline = "PLACEHOLDER"
    for cat in categories:
        percentage = ""
        count = 0
        while grade_dict[cat]['weight'][count].isnumeric() or grade_dict[cat]['weight'][count].isnumeric() == ".":
            percentage.append(grade_dict[cat]['weight'][count])
            count += 1
        
        percentage = float(percentage)
        each = percentage/grade_dict[cat]['number']
        for i in range(grade_dict[cat]['number']):
            worksheet['A' + str(row)] = title
            worksheet['B' + str(row)] = cat + " " + str(i)
            worksheet['C' + str(row)] = deadline
            worksheet['D' + str(row)] = cat
            worksheet['E' + str(row)] = each
        
    






    worksheet.cell(row=3, column=1, value='Jane Smith')
    worksheet.cell(row=3, column=2, value=25)
    workbook.save('example.xlsx')


write_calc(0)